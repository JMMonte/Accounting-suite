# util.py
"""
This file contains the utility functions for the app.

Functions:
- random_fill_days: Fill business days based on max_total constraint,
working with the Excel template's percentage-based calculation system.
- split_dia_hora: Split a date and time string into two parts.
- export_to_excel: Export the preview DataFrame to an Excel file.
- group_consecutive_days: Group consecutive days into trips.
- categorize_trips: Categorize trips into 100%, 75%, 50%, and 25% values.
"""
# Standard library imports
import random
from io import BytesIO
from datetime import datetime
import io

# Third party imports
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

# Local imports
from constants import (
    EXCEL_START_ROW,
    EXCEL_MAX_ROW,
    EXCEL_COLUMN_MAPPING,
)


def _simulate_excel_calculation(filled_days, max_daily, optimize_for_target=False):
    """
    Simulate the Excel calculation to estimate the total.
    Simplified to match the new categorize_trips logic.
    """
    if not filled_days:
        return 0

    # Group into trips to simulate categorize_trips behavior
    filled_sorted = sorted(filled_days, key=lambda x: x["Data"])

    # Add date objects for grouping
    for d in filled_sorted:
        d["_date_obj"] = datetime.strptime(d["Data"], "%Y-%m-%d").date()

    # Group consecutive days
    trips = []
    if filled_sorted:
        trip = [filled_sorted[0]]
        for prev, curr in zip(filled_sorted, filled_sorted[1:]):
            if (curr["_date_obj"] - prev["_date_obj"]).days == 1:
                trip.append(curr)
            else:
                trips.append(trip)
                trip = [curr]
        trips.append(trip)

    # Count percentages as Excel would
    count_100 = count_75 = count_50 = count_25 = 0

    for trip in trips:
        n = len(trip)
        for i, d in enumerate(trip):
            if n == 1:
                # Single day trips are always 100%
                count_100 += 1
            else:
                if i == 0:
                    if optimize_for_target:
                        # Expected values: 80% chance of 100%, 20% chance of 75%
                        count_100 += 0.8
                        count_75 += 0.2
                    else:
                        # First day - mostly 100%
                        if random.random() < 0.8:
                            count_100 += 1
                        else:
                            count_75 += 1
                elif i == n - 1:
                    if optimize_for_target:
                        # Expected values: 70% chance of 50%, 30% chance of 25%
                        count_50 += 0.7
                        count_25 += 0.3
                    else:
                        # Last day - no 0% days
                        if random.random() < 0.7:
                            count_50 += 1
                        else:
                            count_25 += 1
                else:
                    # Middle days are always 100%
                    count_100 += 1

    # Clean up temporary date objects
    for d in filled_sorted:
        if "_date_obj" in d:
            del d["_date_obj"]

    # Calculate total as Excel does
    total = (
        (count_100 * max_daily * 1.0)
        + (count_75 * max_daily * 0.75)
        + (count_50 * max_daily * 0.50)
        + (count_25 * max_daily * 0.25)
    )
    return round(total, 2)


def random_fill_days(business_days, max_daily, max_total):
    """
    Fill business days based on max_total constraint using an iterative approach.
    First fills with high-value days, then adds partial days to maximize budget usage.
    """
    days = business_days.copy()
    random.shuffle(days)

    # Phase 1: Start with a conservative number of days
    # Assume average of ~90% per day to be safe
    estimated_avg = 0.90
    initial_days = int(max_total / (max_daily * estimated_avg))
    initial_days = min(initial_days, len(days))

    # Find a good starting point
    best_config = []
    best_total = 0

    # Try different initial configurations
    for num_days in range(max(1, initial_days - 2), min(len(days), initial_days + 3)):
        test_days = days[:num_days]
        test_filled = []
        for d in sorted(test_days):
            test_filled.append(
                {
                    "Dia": d.day,
                    "Data": d.strftime("%Y-%m-%d"),
                    "Dia da Semana": d.strftime("%A"),
                    "Valor (€)": max_daily,
                }
            )

        # Calculate expected total
        expected_total = _simulate_excel_calculation(
            test_filled, max_daily, optimize_for_target=True
        )

        if expected_total <= max_total and expected_total > best_total:
            best_config = test_days.copy()
            best_total = expected_total

    # Phase 2: Iteratively add more days to use remaining budget
    # This creates opportunities for 25% and 50% days
    remaining_days = [d for d in days if d not in best_config]

    while (
        remaining_days and best_total < max_total * 0.95
    ):  # Try to get within 95% of target
        # Find consecutive pairs in remaining days (these create 25% days efficiently)
        consecutive_pairs = []
        remaining_sorted = sorted(remaining_days)

        i = 0
        while i < len(remaining_sorted) - 1:
            if (remaining_sorted[i + 1] - remaining_sorted[i]).days == 1:
                consecutive_pairs.append((remaining_sorted[i], remaining_sorted[i + 1]))
                i += 2
            else:
                i += 1

        # Try adding a consecutive pair first (creates a 25% day)
        added_something = False
        for pair in consecutive_pairs:
            test_config = best_config + list(pair)
            test_filled = []
            for d in sorted(test_config):
                test_filled.append(
                    {
                        "Dia": d.day,
                        "Data": d.strftime("%Y-%m-%d"),
                        "Dia da Semana": d.strftime("%A"),
                        "Valor (€)": max_daily,
                    }
                )

            # Check if this stays under budget
            worst_case = 0
            for _ in range(5):
                simulated = _simulate_excel_calculation(
                    test_filled, max_daily, optimize_for_target=False
                )
                worst_case = max(worst_case, simulated)

            if worst_case <= max_total:
                best_config = test_config
                best_total = _simulate_excel_calculation(
                    test_filled, max_daily, optimize_for_target=True
                )
                for d in pair:
                    remaining_days.remove(d)
                added_something = True
                break

        # If no pairs work, try single days
        if not added_something and remaining_days:
            for day in remaining_days[:5]:  # Try first 5 remaining days
                test_config = best_config + [day]
                test_filled = []
                for d in sorted(test_config):
                    test_filled.append(
                        {
                            "Dia": d.day,
                            "Data": d.strftime("%Y-%m-%d"),
                            "Dia da Semana": d.strftime("%A"),
                            "Valor (€)": max_daily,
                        }
                    )

                # Check if this stays under budget
                worst_case = 0
                for _ in range(5):
                    simulated = _simulate_excel_calculation(
                        test_filled, max_daily, optimize_for_target=False
                    )
                    worst_case = max(worst_case, simulated)

                if worst_case <= max_total:
                    best_config = test_config
                    best_total = _simulate_excel_calculation(
                        test_filled, max_daily, optimize_for_target=True
                    )
                    remaining_days.remove(day)
                    added_something = True
                    break

        if not added_something:
            break  # Can't add any more days

    # Prepare final result
    filled_days_list = []
    for d in sorted(best_config):
        filled_days_list.append(
            {
                "Dia": d.day,
                "Data": d.strftime("%Y-%m-%d"),
                "Dia da Semana": d.strftime("%A"),
                "Valor (€)": max_daily,
            }
        )

    return filled_days_list, best_total


def split_dia_hora(val):
    """
    Split a date and time string into two parts.
    """
    if val and isinstance(val, str) and len(val.split()) == 2:
        d, h = val.split()
        try:
            d = datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError:
            pass
        return d, h
    return "", ""


def export_to_excel(
    template_path,
    preview_df,
    company_name,
    nipc,
    company_address,
    gestor_name,
    gestor_address,
    gestor_nifps,
    gestor_categoria,
    signature_file=None,
    max_daily=None,
):
    """
    Export the preview DataFrame to an Excel file.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    ws["C4"] = company_name
    ws["C5"] = nipc
    # Add company address to merged cell C6-I6
    ws["C6"] = company_address
    # Add Gestor info
    ws["C40"] = gestor_name
    ws["C41"] = gestor_address
    ws["C42"] = gestor_nifps
    ws["C43"] = gestor_categoria
    # Set Valor Atribuído (max_daily) in I40
    if max_daily is not None:
        ws["I40"] = max_daily
    # Add signature image (uploaded or default)

    if signature_file is not None:
        image_bytes = signature_file.read()
        pil_img = PILImage.open(io.BytesIO(image_bytes))
    else:
        pil_img = PILImage.open("black.png")
    # Resize to max width 200px, keep aspect ratio, don't upscale
    max_width = 200
    w, h = pil_img.size
    if w > max_width:
        new_w = max_width
        new_h = int(h * (max_width / w))
        pil_img = pil_img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)
    img_buffer = io.BytesIO()
    pil_img.save(img_buffer, format="PNG")
    img_buffer.seek(0)
    xl_img = XLImage(img_buffer)
    ws.add_image(xl_img, "C65")
    start_row = EXCEL_START_ROW
    for i, row in enumerate(preview_df.iterrows(), start=start_row):
        if i > EXCEL_MAX_ROW:
            break
        row_dict = row[1].to_dict()
        mapping = EXCEL_COLUMN_MAPPING
        for col, key in mapping:
            value = row_dict.get(key, "")
            cell = ws[f"{col}{i}"]
            if not isinstance(cell, MergedCell):
                cell.value = value
        for col, col_name in zip(["K", "L", "M", "N"], ["100%", "75%", "50%", "25%"]):
            cell = ws[f"{col}{i}"]
            if not isinstance(cell, MergedCell):
                value = row_dict.get(col_name, 0)
                cell.value = value
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def group_consecutive_days(filled_days):
    """
    Group consecutive days into trips.
    """
    filled_sorted = sorted([d for d in filled_days], key=lambda x: x["Data"])
    trips = []
    if filled_sorted:
        for d in filled_sorted:
            d["_date_obj"] = datetime.strptime(d["Data"], "%Y-%m-%d").date()
        trip = [filled_sorted[0]]
        for prev, curr in zip(filled_sorted, filled_sorted[1:]):
            if (curr["_date_obj"] - prev["_date_obj"]).days == 1:
                trip.append(curr)
            else:
                trips.append(trip)
                trip = [curr]
        trips.append(trip)
        for d in filled_sorted:
            del d["_date_obj"]
    return trips


def categorize_trips(trips, OBJECTIVES, CLIENT_ADDRESS):
    """
    Categorizes trips and assigns percentage values based on trip duration and times.
    Simplified implementation for reliability and reaching target values.

    Strategy for optimization:
    - Single day trips: Always 100% (full day at client)
    - Multi-day trips:
      - First day: Mostly 100% (early departure), sometimes 75% (afternoon)
      - Middle days: Always 100%
      - Last day: Mostly 50% (evening return), sometimes 25% (afternoon)

    This ensures we maximize value while being realistic.

    Args:
        trips: List of trips, where each trip is a list of consecutive filled days
        OBJECTIVES: List of possible objectives for the trips
        CLIENT_ADDRESS: The client's address to use for trip locations

    Returns:
        List of all individual days with categorization applied
    """
    all_categorized_days = []
    trip_id = 0

    for trip in trips:
        trip_length = len(trip)

        # Apply the same objective to all days in the trip
        objetivo = random.choice(OBJECTIVES)
        local = CLIENT_ADDRESS

        for i, day in enumerate(trip):
            # Reset all percentage columns
            day["Valor 100% (€)"] = ""
            day["Valor 75% (€)"] = ""
            day["Valor 50% (€)"] = ""
            day["Valor 25% (€)"] = ""

            # Determine the percentage based on trip length and position
            if trip_length == 1:
                # Single day trip - always 100%
                day["Valor 100% (€)"] = 1
                inicio_dia_hora = f"{day['Data']} 07:00"
                regresso_dia_hora = f"{day['Data']} 22:00"
            else:
                if i == 0:
                    # First day of multi-day trip
                    # Mostly early departure (100%), sometimes afternoon (75%)
                    if random.random() < 0.8:  # 80% chance
                        day["Valor 100% (€)"] = 1
                        inicio_hora = "08:00"
                    else:  # 20% chance
                        day["Valor 75% (€)"] = 1
                        inicio_hora = "14:00"

                    inicio_dia_hora = f"{day['Data']} {inicio_hora}"
                    regresso_dia_hora = f"{day['Data']} 23:59"

                elif i == trip_length - 1:
                    # Last day of multi-day trip
                    # Always return in afternoon or evening (never morning for 0%)
                    if random.random() < 0.7:  # 70% chance
                        day["Valor 50% (€)"] = 1
                        regresso_hora = "21:00"
                    else:  # 30% chance
                        day["Valor 25% (€)"] = 1
                        regresso_hora = "18:00"

                    inicio_dia_hora = f"{day['Data']} 08:00"
                    regresso_dia_hora = f"{day['Data']} {regresso_hora}"

                else:
                    # Middle days - always 100%
                    day["Valor 100% (€)"] = 1
                    inicio_dia_hora = f"{day['Data']} 08:00"
                    regresso_dia_hora = f"{day['Data']} 23:59"

            # Add the additional required fields
            day.update(
                {
                    "inicio (Dia Hora)": inicio_dia_hora,
                    "regresso (Dia Hora)": regresso_dia_hora,
                    "Mapa deslocação / Objectivo": objetivo,
                    "Local onde foram prestados": local,
                    "_trip_id": trip_id,
                }
            )

        all_categorized_days.extend(trip)
        trip_id += 1

    return all_categorized_days
